import openpyxl


def estatistica():
    try:
        # Abre a planilha
        wb = openpyxl.load_workbook("dados.xlsx")

        # Verifica se a aba "lucro" existe, se não cria
        if "lucro" not in wb.sheetnames:
            wb.create_sheet("lucro")

        sheet = wb["lucro"]

        # Define as células iniciais e finais dos dados (ignorando o cabeçalho)
        primeiraLinha = 2
        ultimaLinha = sheet.max_row

        # Inicializa as variáveis para os totais
        custoTotalTotal = 0
        lucroTotal = 0
        precoDeVendaTotal = 0

        # Calcula os totais
        for linha in range(primeiraLinha, ultimaLinha + 1):
            custoTotalTotal += sheet.cell(row=linha, column=7).value or 0  # Trata células vazias
            lucroTotal += sheet.cell(row=linha, column=8).value or 0
            precoDeVendaTotal += sheet.cell(row=linha, column=3).value or 0

        # Calcula a margem de lucro total, tratando a divisão por zero
        if precoDeVendaTotal != 0:
            margemDeLucroTotal = (lucroTotal / precoDeVendaTotal) * 100
        else:
            margemDeLucroTotal = 0  # Ou "N/A"

        # Formata os valores para duas casas decimais
        custoTotalTotal = round(custoTotalTotal, 2)
        lucroLiquidoTotal = round(lucroTotal, 2)
        margemDeLucroTotal = round(margemDeLucroTotal, 2)

        return [custoTotalTotal, lucroLiquidoTotal, margemDeLucroTotal]

    except FileNotFoundError:
        # Retorna valores padrão se o arquivo não for encontrado
        return [0, 0, 0]
    except Exception as e:
        # Retorna valores padrão para outros erros e imprime o erro no console
        print(f"Erro ao calcular estatísticas: {e}")
        return [0, 0, 0]

def obterDadosExcel(nomeDoArquivo):
    wb = openpyxl.load_workbook(nomeDoArquivo)
    sheet = wb.active
    dados = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        dados.append(row)
    return dados

'''def calcularLucro(nomeProduto, precoDeCompra, precoDeVenda, quantidade, custosAdicionais, custoFrete):
    # Perguntado ao usuário pelos dados
    nomeProduto = nomeProduto
    precoDeCompra = float(precoDeCompra)
    precoDeVenda = float(precoDeVenda)
    quantidade = int(quantidade)
    custosAdicionais = float(custosAdicionais)
    custoFrete = float(custoFrete)

    # Calculando o lucro
    custoTotal = (precoDeCompra + custosAdicionais) + custoFrete * quantidade
    lucro = (precoDeVenda - precoDeCompra - custosAdicionais - custoFrete) * quantidade
    margemDeLucro = lucro / (precoDeVenda * quantidade) * 100

    # Imprimindo os resultados

    print("O lucro do produto {} é de R$ {:.2f} e a margem de lucro é de {:.2f}%.".format(nomeProduto, lucro, margemDeLucro))
    print("O cisto total do produto {} é de R$ {:.2f}.".format(nomeProduto, custoTotal))'''

def salvarProdutos(nomeProduto, precoDeCompra, precoDeVenda, quantidade, custosAdicionais, custoFrete):
    # Converte os valores para os tipos corretos, tratando possíveis erros de conversão
    try:
        precoDeCompra = float(precoDeCompra)
        precoDeVenda = float(precoDeVenda)
        quantidade = int(quantidade)
        custosAdicionais = float(custosAdicionais)
        custoFrete = float(custoFrete)
    except ValueError:
        # Se a conversão falhar, retorna um erro
        print("Erro: Valores inválidos inseridos. Certifique-se de que os valores sejam numéricos.")
        return

    # Calcula o lucro
    custoTotal = (precoDeCompra * quantidade) + custosAdicionais + (custoFrete * quantidade)
    lucro = (precoDeVenda - precoDeCompra - custosAdicionais - custoFrete) * quantidade
    margemDeLucro = (lucro / (precoDeVenda * quantidade)) * 100

    try:
        wb = openpyxl.load_workbook("dados.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Verificar se a planilha "lucro" existe, senão criar
    if "lucro" not in wb.sheetnames:
        wb.create_sheet("lucro")

    sheet = wb["lucro"]

    # Definir cabeçalho se a planilha for nova
    if sheet.max_row == 1:
        sheet.append(["Produto", "Preço de Compra", "Preço de Venda", "Quantidade", "Custos Adicionais", "Custo de Frete", "Custo Total", "Lucro Líquido", "Margem de Lucro"])

    # Salva os resultados na planilha
    sheet.append([nomeProduto, precoDeCompra, precoDeVenda, quantidade, custosAdicionais, custoFrete, custoTotal, lucro, margemDeLucro])

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save("dados.xlsx")
    print("Produto salvo com sucesso!")

def deletarLinhaPorNome(nomeProduto, nomePlanilha):
    # Carrega a planilha
    wb = openpyxl.load_workbook(nomePlanilha)
    sheet = wb.active
    contador = 2

    # Percorre a coluna de nomes procurando pelo nome do produto
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if str(row[0]) == nomeProduto:
            # Obtém o número da linha e deleta a linha inteira
            linha = contador
            sheet.delete_rows(linha)
            break

        # Incrementado o contador
        contador += 1

    # Salva as mudanças na planilha
    wb.save(nomePlanilha)
