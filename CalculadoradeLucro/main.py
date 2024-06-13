import openpyxl

def calcularLucro(nomeProduto, precoDeCompra, precoDeVenda, quantidade, custosAdicionais, custoFrete):
    # Perguntado ao usuário pelos dados
    nomeProduto = nomeProduto
    precoDeCompra = float(precoDeCompra)
    precoDeVenda = float(precoDeVenda)
    quantidade = int(quantidade)
    custosAdicionais = float(custosAdicionais)
    custoFrete = float(custoFrete)

    # Calculando o lucro
    custoTotal = (precoDeCompra + custosAdicionais) + custoFrete * quantidade
    lucro = (precoDeVenda - precoDeCompra - custosAdicionais - custoFrete) * quantidade
    margemDeLucro = lucro / (precoDeVenda * quantidade) * 100

    # Imprimindo os resultados

    print("O lucro do produto {} é de R$ {:.2f} e a margem de lucro é de {:.2f}%.".format(nomeProduto, lucro, margemDeLucro))
    print("O cisto total do produto {} é de R$ {:.2f}.".format(nomeProduto, custoTotal))

    # Salvando os resultados em um arquivo.xlsx

    try:
        wb = openpyxl.load_workbook("dados.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.active.title  = "Resultado da Calculadora de Lucro"
        wb.active.append(["Produto", "Preço de Compra", "Preço de Venda", "Quantidade", "Custos Adicionais", "Custo de Frete", "Custo Total", "Lucro liquido", "Margem de lucro"])

    sheet = wb["Lucro"]
    lastRow = sheet.max_row + 1

    sheet.cell(row=lastRow, column=1).value = nomeProduto
    sheet.cell(row=lastRow, column=2).value = precoDeCompra
    sheet.cell(row=lastRow, column=3).value = precoDeVenda
    sheet.cell(row=lastRow, column=4).value = quantidade
    sheet.cell(row=lastRow, column=5).value = custosAdicionais
    sheet.cell(row=lastRow, column=6).value = custoFrete
    sheet.cell(row=lastRow, column=7).value = custoTotal
    sheet.cell(row=lastRow, column=8).value = lucro
    sheet.cell(row=lastRow, column=9).value = margemDeLucro

    wb.save("dados.xlsx")

